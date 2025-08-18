# -*- coding: utf-8 -*-
import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import REPORT_DIR, CPU_THRESHOLD, MEM_THRESHOLD, DISK_THRESHOLD

def set_column_auto_width(ws):
    """自动调整列宽"""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

def generate_excel_report(project_name: str, inspector: str, date_str: str, rows: List[Dict[str, Any]]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "日常巡检报告"

    # 样式
    bold_center = Font(bold=True)
    red_font = Font(color="FF0000")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    # 基本信息
    ws["A1"], ws["B1"] = "项目名称", project_name
    ws["A2"], ws["B2"] = "巡检人", inspector
    ws["A3"], ws["B3"] = "巡检时间", date_str
    for r in range(1, 4):
        ws[f"A{r}"].font = bold_center
        ws[f"A{r}"].alignment = center

    # 空一行
    ws.append([])

    # 服务器巡检记录表格
    headers = ["服务器IP", "系统运行时间", "CPU占用情况", "内存使用情况", "磁盘空间占用情况"]
    ws.append(headers)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(ws.max_row, c)
        cell.font = bold_center
        cell.alignment = center
        cell.border = border

    for row in rows:
        data = [
            row.get("ip", ""),
            row.get("uptime", ""),
            f'{row.get("cpu",0)}%',
            f'{row.get("mem",0)}%',
            f'{row.get("disk",0)}%'
        ]
        ws.append(data)
        for c in range(1, len(headers)+1):
            cell = ws.cell(ws.max_row, c)
            cell.alignment = center
            cell.border = border

    # 空一行
    ws.append([])
    ws.append(["异常问题描述"])
    ws.cell(ws.max_row, 1).font = bold_center

    abnormal_lines = []
    for r in rows:
        reasons = []
        if not r.get("ok"):
            reasons.append(f'连接失败: {r.get("error","")}')
        else:
            if r.get("cpu", 0) > CPU_THRESHOLD:
                reasons.append(f'CPU {r["cpu"]}% > {CPU_THRESHOLD}%')
            if r.get("mem", 0) > MEM_THRESHOLD:
                reasons.append(f'内存 {r["mem"]}% > {MEM_THRESHOLD}%')
            if r.get("disk", 0) > DISK_THRESHOLD:
                reasons.append(f'磁盘 {r["disk"]}% > {DISK_THRESHOLD}%')
        if reasons:
            abnormal_lines.append(f'{r.get("ip","")}: ' + "; ".join(reasons))

    if abnormal_lines:
        for line in abnormal_lines:
            ws.append([line])
            ws.cell(ws.max_row, 1).font = red_font
    else:
        ws.append(["无"])

    # 自动调整列宽
    set_column_auto_width(ws)

    # 保存
    os.makedirs(REPORT_DIR, exist_ok=True)
    fname = f"巡检报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(REPORT_DIR, fname)
    wb.save(out_path)
    return out_path
